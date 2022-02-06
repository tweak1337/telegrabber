import time
from pyrogram import Client, filters, emoji
from pyrogram.errors import FloodWait
from pyrogram.types import ChatPermissions

import random
from datetime import date, datetime, timedelta
from time import sleep
import pymorphy2
import psycopg2
from sqlalchemy import create_engine
import openpyxl
import random
import time
import os

app = Client('my_caccount2')

m = pymorphy2.MorphAnalyzer()


# функция нормализации текста для каждого слова к первое склонение ед.число
def normalize(s1):
    sent = []
    stroka = s1.replace('?', '')
    stroka = stroka.replace('!', '')
    stroka = stroka.replace('.', '')
    stroka = stroka.replace(',', '')

    s = stroka.lower().split()

    for i in s:
        p = m.parse(i)[0]
        sent.append(p.normal_form)
    s2 = ' '.join(sent)

    return s2


with open('config.ini', 'r') as f:
    data = f.read().splitlines()
    donor = data[3].split(' = ')[1]
    own = data[4].split(' = ')[1]


def connection():

    dbpass = os.getenv('dbpass')
    dbname = os.getenv('dbname')
    dbuser = os.getenv('dbuser')
    dbhost = os.getenv('dbhost')
    dbport = os.getenv('dbport')

    con = psycopg2.connect(database='domgrabber', user = 'postgres', password = 'Rftgyhvbn_4756',
                           host = 'localhost', port = '5432')

    cursor = con.cursor()

    engine = create_engine(f"postgresql+psycopg2://{dbuser}:{dbpass}@{dbhost}:{dbport}/{dbname}?charset=utf8mb4", echo = False)

    return con, cursor, engine

# Фильтры
def fillta():

    x = []
    y = []
    file =r'Z:\pythonprojects\domrf_grabber 3.0\Фильтры.xlsx'
    wb = openpyxl.load_workbook(file, data_only=True)
    wb.active = 0
    ws = wb.active
    first_filters_list = ws['A2':'A'+str(ws.max_row)]
    second_filters_list = ws['B2':'B'+str(ws.max_row)]

    for row in first_filters_list:
        for cell in row:
            if cell.value is None:
                break
            else:
                x.append(str(cell.value).lower())

    for row in second_filters_list:
        for cell in row:
            if cell.value is None:
                break
            else:
                y.append(str(cell.value).lower())

    # функция нормализации текста для каждого слова к первое склонение ед.число
    def normalize1(s1):
        sent = []
        stroka = s1.replace('?', '')
        stroka = stroka.replace('!', '')
        stroka = stroka.replace('.','')
        stroka = stroka.replace(',','')

        s = stroka.lower().split()

        for i in s:
            p = m.parse(i)[0]
            sent.append(p.normal_form)
        s2 = ' '.join(sent)

        return s2

    lenx = len(x)
    leny = len(y)

    # Создаем список из фильтров, в списке у заменяем пробелы на -, чтобы слить

    total_list = []
    for i in range(0, lenx):
        for j in range(0, leny):
            total_list.append(x[i] +' '+ y[j].replace(' ', '-'))


    lent = len(total_list)

    # По пробелам разделяем слова на отдельные элементы списка
    global_list = []
    for i in range(0,lent):
        global_list.append(total_list[i].split())

    # Обратно заменяем - на пробелы
    for i in range(len(global_list)):
        for j in range(len(global_list[i])):
            global_list[i][j] = normalize1(global_list[i][j].replace('-', ' '))

    # Выбираем позитивные и негативные слова
    wb.active = 1
    ws = wb.active
    cola = ws['A2':'A' + str(ws.max_row)]
    colb = ws['B2':'B' + str(ws.max_row)]
    cole = ws['E2':'E' + str(ws.max_row)]
    cold = ws['D2':'D' + str(ws.max_row)]
    negatives = []
    for row in cola:
        for cell in row:
            negatives.append(cell.value)

    positives = []
    for row in colb:
        for cell in row:
            positives.append(cell.value)
    positive_smiles0 = []
    for row in cold:
        for cell in row:
            positive_smiles0.append(cell.value)

    negative_smiles0 = []
    for row in cole:
        for cell in row:
            negative_smiles0.append(cell.value)

    return global_list, negatives, positives, positive_smiles0, negative_smiles0

global_list, negatives, positives, positive_smiles0, negative_smiles0 = fillta()
# global_list = [hui.lower() for hui in global_list]

positive_smiles = []
negative_smiles = []
negatives_list = []
positives_list = []

# Убираем None, переводим в нижний регистр
for i in negatives:
    if i == None:
        pass
    else:
        negatives_list.append(normalize(i.lower()))

for i in positives:
    if i == None:
        pass
    else:
        positives_list.append(normalize(i.lower()))

for i in positive_smiles0:
    if i != None:
        positive_smiles.append(i)

for i in negative_smiles0:
    if i != None:
        negative_smiles.append(i)

app.start()
donor = ['ria_realty', 'fontankaspb', 'nedvizha', 'realestate_rf', 'Jelezobetonniyzames', 'banksta', 'ruarbitr',
         'posadky', 'moscowtop', 'domostroy_channel', 'vchkogpu', 'propertyinsider', 'filatofff', 'cotlovan_contrust',
         'belaya_kaska', 'Mos_stroi', 'domtech', 'lietomerealty', 'atsogoev', 'novostroyman', 'me',
         'trubapodneglinnoy', 'riskovik', 'wearestroyka', 'soyzsmet', 'riskovik2', 'real_estate', 'everydayproperty',
         'kompr', 'zmysl', 'chpmoscow', 'zarbitrazhy', 'kleveta']

def mission():
    for public in donor:
        print(public)
        for message in app.iter_history(public, limit= 1):

            posted_timestamp = datetime.utcfromtimestamp(message.date)
            posted_timestamp = posted_timestamp + timedelta(hours=3)

            now = datetime.now()
            delta = now - timedelta(days = 2)


            message_id = message.message_id
            username = message.chat.username
            channel_name = message.chat.title
            try:
                try:
                    replied_from = message.forward_from_chat.username
                    replied_from_name = message.forward_from_chat.title
                except Exception as e:
                    replied_from = message.forward_from.username
                    replied_from_name = message.forward_from.first_name
            except Exception as e:
                replied_from = None
                replied_from_name = None

            try:
                if posted_timestamp < delta:
                    lviews = int(message.views)
                else:
                    lviews = None

            except Exception:

                lviews = None

            con, cursor, engine = connection()
            now = datetime.now()
            now = now.strftime("%Y-%m-%d %H:%M:%S")

            if not message.text is None:
                finaltext = message.text
            elif not message.caption is None:
                finaltext = message.caption
            elif not message.video is None and message.caption is None and message.text is None:
                finaltext = 'No caption video'
            else:
                finaltext = 'No caption image'

            text = normalize(finaltext)
            txt = text.lower()
            txt0 = txt.split()
            txt1 = finaltext.split()
            str_hashtags = None

            if '#' in text:
                hashtagsdb = []
                str_hashtags = ''

                for i in txt1:
                    if '#' in i:
                        hashtagsdb.append(i)
                for i in hashtagsdb:
                    if i != hashtagsdb[-1]:
                        str_hashtags += str(i) + ', '
                    else:
                        str_hashtags += str(i)

            # Считаем позитивные и негативные слова
            bad = 0
            good = 0
            good_smiles = 0
            bad_smiles = 0

            cursor.execute("select message_id from dom.dom_rf where message_id = %s and username = %s", (message_id, username))
            if not cursor.fetchall():

                # Cчитаем негативн и позитив
                # Smiles
                for pos_smile in positive_smiles:
                    for letter in finaltext:
                        if letter == pos_smile:
                            good_smiles += 1

                for neg_smile in negative_smiles:
                    for letter in finaltext:
                        if letter == neg_smile:
                            bad_smiles += 1

                for nega in negatives_list:
                    for word in txt0:
                        if nega == word:
                            bad += 1

                for pos in positives_list:
                    for word in txt0:
                        if pos == word:
                            good += 1

                if good_smiles > 0 or bad_smiles > 0:
                    if good_smiles > 0 and bad_smiles == 0:
                        emotional = 'Positive'
                    elif bad_smiles > 0 and good_smiles == 0:
                        emotional = 'Negative'
                    elif good_smiles / (good_smiles + bad_smiles) < 0.3:
                        emotional = 'Negative'
                    elif bad_smiles / (good_smiles + bad_smiles) < 0.3:
                        emotional = 'Positive'
                    else:
                        if bad == 0 and good == 0:
                            emotional = 'Neutral'
                        elif bad > 0 and good == 0:
                            emotional = 'Negative'
                        elif good > 0 and bad == 0:
                            emotional = 'Positive'
                        elif good / (good + bad) < 0.5:
                            emotional = 'Negative'
                        elif bad / (good + bad) < 0.3:
                            emotional = 'Positive'
                        else:
                            emotional = "Undefined"
                else:
                    if bad == 0 and good == 0:
                        emotional = 'Neutral'
                    elif bad > 0 and good == 0:
                        emotional = 'Negative'
                    elif good > 0 and bad == 0:
                        emotional = 'Positive'
                    elif good / (good + bad) < 0.5:
                        emotional = 'Negative'
                    elif bad / (good + bad) < 0.3:
                        emotional = 'Positive'
                    else:
                        emotional = "Undefined"

                # проверка на рекламу
                if ('t.me' in str(message) or '"type": "mention"' in str(message) or
                        '"type": "text_link"' in str(message) or 'clc.to' in str(message) or
                        '#реклама' in str(message) or 'InlineKeyboardMarkup' in str(message)):
                    ad = 1
                else:
                    ad = 0

                cursor.execute(
                    '''insert into dom.dom_rf (message_id, username, channel_name, replied_from, replied_from_name,
                     message_text, posted_timestamp, activity_date, emotional, is_ad, hashtags, lviews) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                    (message_id, username, channel_name, replied_from, replied_from_name, finaltext, posted_timestamp, now, emotional, ad,
                     str_hashtags,lviews))
                con.commit()



            else:  # Если запись уже есть в БД, то обновляем ее, обновляя сообщение, добавляя время изменения, изменяя счетчик изменений
                pass
mission()
        # cursor.execute("select edit_counter from dom_rf where message_id = %s", message_id)
        # edit_counter = cursor.fetchall()
        # edit_counter = list(edit_counter[0].values())[0]
        # if edit_counter is None:
        #
        #     cursor.execute(
        #         "UPDATE dom_rf SET message_text=%s, edited_timestamp = %s, edit_counter = %s WHERE message_id = %s",
        #         (finaltext, now, 1, message_id))
        #
        # else:
        #
        #     cursor.execute(
        #         "UPDATE dom_rf SET message_text=%s, edited_timestamp = %s, edit_counter = %s WHERE message_id = %s",
        #         (finaltext, now, edit_counter + 1, message_id))
        #
        # con.commit()
