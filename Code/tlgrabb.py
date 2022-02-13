import time
from pyrogram import Client, filters, emoji
from pyrogram.errors import FloodWait
from pyrogram.types import ChatPermissions
import random
from datetime import date, datetime, timedelta
from time import sleep
import pymorphy2
import psycopg2
from sqlalchemy import create_engine
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from xlsxwriter import Workbook


app = Client('my_account')


m = pymorphy2.MorphAnalyzer()
# функция нормализации текста для каждого слова к первое склонение ед.число
def normalize(s1):
    sent = []
    stroka = s1.replace('?', '')
    stroka = stroka.replace('!', '')
    stroka = stroka.replace('.','')
    stroka = stroka.replace(',','')

    s = stroka.lower().split()

    for i in s:
        p = m.parse(i)[0]
        sent.append(p.normal_form)
    s2 = ' '.join(sent)

    return s2



with open ('config.ini','r') as f:
    data = f.read().splitlines()
    donor = data[3].split(' = ')[1]
    own = data[4].split(' = ')[1]



# Фильтры
def fillta():

    x = []
    y = []
    file =r'Z:\pythonprojects\telegrabber 3.0\filters.xlsx'
    wb = openpyxl.load_workbook(file, data_only=True)


    # Выбираем позитивные и негативные слова и смайлы
    wb.active = 0
    ws = wb.active
    cola = ws['A2':'A' + str(ws.max_row)]
    colb = ws['B2':'B' + str(ws.max_row)]
    cold = ws['D2':'D' + str(ws.max_row)]
    cole = ws['E2':'E' + str(ws.max_row)]
    negatives = []
    for row in cola:
        for cell in row:
            negatives.append(cell.value)

    positives = []
    for row in colb:
        for cell in row:
            positives.append(cell.value)

    positive_smiles0 = []
    for row in cold:
        for cell in row:
            positive_smiles0.append(cell.value)

    negative_smiles0 = []
    for row in cole:
        for cell in row:
            negative_smiles0.append(cell.value)

    return negatives, positives, positive_smiles0, negative_smiles0

negatives, positives, positive_smiles0, negative_smiles0 = fillta()
# global_list = [hui.lower() for hui in global_list]
negatives_list = []
positives_list = []
positive_smiles = []
negative_smiles = []

# Убираем None, переводим в нижний регистр
for i in negatives:
    if i == None:
        pass
    else:
        negatives_list.append(normalize(i.lower()))

for i in positives:
    if i == None:
        pass
    else:
        positives_list.append(normalize(i.lower()))

for i in positive_smiles0:
    if i != None:
        positive_smiles.append(i)

for i in negative_smiles0:
    if i != None:
        negative_smiles.append(i)

def connection():

    with open('dbconn.ini', 'r') as z:
        all = z.read().splitlines()
        database = all[0].split(' = ')[1]
        dbuser = all[1].split(' = ')[1]
        dbpass = all[2].split(' = ')[1]
        dbhost = all[3].split(' = ')[1]
        dbport = all[4].split(' = ')[1]

    con = psycopg2.connect(database=database, user = dbuser, password = dbpass,
                           host = dbhost, port = dbport)

    cursor = con.cursor()

    engine = create_engine("postgresql+psycopg2://root:Rftgyhvbn_4756@localhost/tested?charset=utf8mb4", echo = False)

    return con, cursor, engine


@app.on_message(filters.chat(eval(donor)))
async def get_post(client, message):

    #Создаем все переменные для бд
    message_id = message.message_id
    username = message.chat.username
    channel_name = message.chat.title
    link = f'https://t.me/{username}/{message_id}'

    try:
        webpage_url = message.web_page.url
    except Exception:
        webpage_url = None

    try:
        caption_entities = message.caption_entities
    except Exception:
        caption_entities = None

    try:
        entities = message.entities
    except Exception:
        entities = None

    try:
        try:
            replied_from = message.forward_from_chat.username
            replied_from_name = message.forward_from_chat.title
        except Exception as e:
            replied_from = message.forward_from.username
            replied_from_name = message.forward_from.first_name
    except Exception as e:
        replied_from = None
        replied_from_name = None

    try:
        lviews = await client.get_history(username)
        lviews = lviews[30]
        mesid50 = lviews.message_id
        lviews = int(lviews.views)

    except Exception:
        mesid50 = None
        lviews = None

    posted_timestamp = datetime.utcfromtimestamp(message.date)
    posted_timestamp = posted_timestamp + timedelta(hours=3)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    con, cursor, engine = connection()

    hashtags = ['#videos', '#unexpected', '#nextfuckinglevel', '#funny', '#wellthatsucks', '#interestingasfuck']

    if not message.text is None:
        finaltext = message.text
    elif not message.caption is None:
        finaltext = message.caption
    elif not message.video is None and message.caption is None and message.text is None:
        finaltext = 'No caption video'
    else:
        finaltext = 'No caption image'
    # print(message)
    # Проверка на рекламу
    if ('clc.to' in str(message) or '#реклама' in str(message) or 'InlineKeyboardMarkup' in str(message)):
        ad = 1
        # print(now,'0')
    elif '"type": "mention"' in str(message):
        if str(username).lower() not in str(finaltext).lower():
            # print(now,'1')
            ad = 1
        else:
            # print(now,'2')
            ad = 0
    elif webpage_url != None:

        if str(username).lower() not in webpage_url.lower():
            # print(now,'3')
            ad = 1
        else:
            # print(now,'4')
            ad = 0

    elif 't.me' in str(message) and caption_entities != None:

        if str(username).lower() not in str(caption_entities).lower():
            # print(now,'5')
            ad = 1
        else:
            # print(now,'6')
            ad = 0

    elif 't.me' in str(message) and entities != None:

        if str(username).lower() not in str(entities).lower():
            # print(now,7)
            ad = 1
        else:
            # print(now,8)
            ad = 0
    else:
        # print(now, 9)
        ad = 0

    text = normalize(finaltext)
    txt = text.lower()
    txt0 = txt.split()
    txt1 = finaltext.split()
    str_hashtags = None
    if '#' in text:
        hashtagsdb = []
        str_hashtags = ''

        for i in txt1:
            if '#' in i:
                hashtagsdb.append(i)
        for i in hashtagsdb:
            if i != hashtagsdb[-1]:
                str_hashtags += str(i) + ', '
            else:
                str_hashtags += str(i)


    # Устанавливаем счетчик количества просмотров поста 30 сообщений назад
    cursor.execute("UPDATE telegram.telegrabber SET lviews = %s where message_id = %s and username = %s", (lviews, mesid50, username))
    con.commit()

    # Считаем позитивные и негативные слова и смайлы
    bad = 0
    good = 0

    good_smiles = 0
    bad_smiles = 0

    # Счетчик времени для отложенных постов

    today = datetime.today()
    unixtime = time.mktime(today.timetuple())
    randomint = random.randint(60,360)
    # Селектим все id проверяем есть ли такой
    cursor.execute("select message_id from telegram.telegrabber where message_id = %s and username = %s", (message_id, username))
    if not cursor.fetchall():

        # Cчитаем негативн и позитив

        #Smiles
        for pos_smile in positive_smiles:
            for letter in finaltext:
                if letter == pos_smile:
                    good_smiles += 1

        for neg_smile in negative_smiles:
            for letter in finaltext:
                if letter == neg_smile:
                    bad_smiles += 1

        for nega in negatives_list:
            for word in txt0:
                if nega == word:
                    bad += 1

        for pos in positives_list:
            for word in txt0:
                if pos == word:
                    good += 1

        if good_smiles>0 or bad_smiles>0:
            if good_smiles > 0 and bad_smiles == 0:
                emotional = 'Positive'
            elif bad_smiles >0 and good_smiles == 0:
                emotional = 'Negative'
            elif good_smiles / (good_smiles+bad_smiles)< 0.3:
                emotional = 'Negative'
            elif bad_smiles / (good_smiles+bad_smiles)< 0.3:
                emotional = 'Positive'
            else:
                if bad == 0 and good == 0:
                    emotional = 'Neutral'
                elif bad > 0 and good == 0:
                    emotional = 'Negative'
                elif good > 0 and bad == 0:
                    emotional = 'Positive'
                elif good / (good + bad) < 0.5:
                    emotional = 'Negative'
                elif bad / (good + bad) < 0.3:
                    emotional = 'Positive'
                else:
                    emotional = "Undefined"
        else:
            if bad == 0 and good == 0:
                emotional = 'Neutral'
            elif bad > 0 and good == 0:
                emotional = 'Negative'
            elif good > 0 and bad == 0:
                emotional = 'Positive'
            elif good / (good + bad) < 0.5:
                emotional = 'Negative'
            elif bad / (good + bad) < 0.3:
                emotional = 'Positive'
            else:
                emotional = "Undefined"


        cursor.execute(
            '''insert into telegram.telegrabber (message_id, username, channel_name, replied_from, replied_from_name,
             message_text, posted_timestamp, activity_date, emotional, is_ad, hashtags, link) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
            (message_id, username, channel_name, replied_from, replied_from_name, finaltext, posted_timestamp, now, emotional, ad, str_hashtags, link))
        con.commit()

        if ad == 0 and username != 'borsh_tg':
            if username == 'Reddit':
                for i in hashtags:
                    if i in finaltext:
                        await message.copy(eval(own))
                    else:
                        pass
            elif username == 'video_dolboeba' or username == 'community_memy' or \
                    username == 'yumor_video_prikoly' or username == 'cmehov' or username == 'ot_zp_do_zp' or \
                    username == 'fiftypound':
                await message.copy(eval(own))

            else:
                pass
        elif username == 'borsh_tg':
            if ad == 0:
                await message.copy(eval(own), caption='@hardcorefunnews')
            elif ad == 1 and 't.me/millions_on_memes' in str(message) or ad == 1 and 't.me/borsh_tg' in str(message):
                await message.copy(eval(own), caption='@hardcorefunnews')
            else:
                pass


    #Если запись уже есть в бд, то обновляем ее, обновляя сообщение, добавляя время изменения, изменяя счетчик изменений
    else:
        cursor.execute(
            "select message_text from telegram.telegrabber where message_id = %s and username = %s order by activity_date desc limit 1",
            (message_id, username))
        comparison = cursor.fetchall()
        comparison_text = comparison[0][0]


        if comparison_text != finaltext:

            len_diff = len(finaltext) - len(comparison_text)

            try:
                edited_timestamp = datetime.utcfromtimestamp(message.edit_date)
                edited_timestamp = edited_timestamp + timedelta(hours=3)
            except Exception as e:
                edited_timestamp = now


            for pos_smile in positive_smiles:
                for letter in finaltext:
                    if letter == pos_smile:
                        good_smiles += 1

            for neg_smile in negative_smiles:
                for letter in finaltext:
                    if letter == neg_smile:
                        bad_smiles += 1

            for nega in negatives_list:
                for word in txt0:
                    if nega == word:
                        bad += 1

            for pos in positives_list:
                for word in txt0:
                    if pos == word:
                        good += 1

            if good_smiles > 0 or bad_smiles > 0:
                if good_smiles > 0 and bad_smiles == 0:
                    emotional = 'Positive'
                elif bad_smiles > 0 and good_smiles == 0:
                    emotional = 'Negative'
                elif good_smiles / (good_smiles + bad_smiles) < 0.3:
                    emotional = 'Negative'
                elif bad_smiles / (good_smiles + bad_smiles) < 0.3:
                    emotional = 'Positive'
                else:
                    if bad == 0 and good == 0:
                        emotional = 'Neutral'
                    elif bad > 0 and good == 0:
                        emotional = 'Negative'
                    elif good > 0 and bad == 0:
                        emotional = 'Positive'
                    elif good / (good + bad) < 0.5:
                        emotional = 'Negative'
                    elif bad / (good + bad) < 0.3:
                        emotional = 'Positive'
                    else:
                        emotional = "Undefined"
            else:
                if bad == 0 and good == 0:
                    emotional = 'Neutral'
                elif bad > 0 and good == 0:
                    emotional = 'Negative'
                elif good > 0 and bad == 0:
                    emotional = 'Positive'
                elif good / (good + bad) < 0.5:
                    emotional = 'Negative'
                elif bad / (good + bad) < 0.3:
                    emotional = 'Positive'
                else:
                    emotional = "Undefined"

            cursor.execute("select edit_counter from telegram.telegrabber where message_id = %s and username = %s order by activity_date desc limit 1", (message_id, username))

            edit_counter = cursor.fetchall()
            edit_counter = edit_counter[0][0]

            if edit_counter is None:

                cursor.execute(
                    '''insert into telegram.telegrabber (message_id, username, channel_name, replied_from, replied_from_name,
                 message_text, posted_timestamp, edited_timestamp, activity_date, emotional, is_ad, hashtags, edit_counter,len_diff, link) 
                 values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                    (message_id, username, channel_name, replied_from, replied_from_name, finaltext,posted_timestamp, edited_timestamp, now, emotional, ad,
                     str_hashtags, 1, len_diff, link))

            else:

                cursor.execute(
                    '''insert into telegram.telegrabber (message_id, username, channel_name, replied_from, replied_from_name,
                 message_text, posted_timestamp, edited_timestamp, activity_date, emotional, is_ad, hashtags, edit_counter, len_diff, link) 
                 values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                    (message_id, username, channel_name, replied_from, replied_from_name, finaltext, posted_timestamp, edited_timestamp, now, emotional, ad,
                     str_hashtags, edit_counter+1, len_diff, link))

            con.commit()
        else:
            pass

        return



app.run()

# else:


# @app.on_message(filters.chat(-492221565))
# async def send_mess(client, message):
#     txts= normalize(message.text.lower())
#     user = message.from_user.username
#     phone = message.from_user.phone_number
#     first_name = message.from_user.first_name
#     last_name = message.from_user.last_name
#     photo = message.from_user.photo
#
#
#     #
#     for one_word in keys3:
#         if one_word in txts:
#             await client.send_message(-492221565, 'КТО СКАЗАЛ "'+str(one_word.upper()) + '"????')
#
#             time.sleep(1)
#             if user is None:
#
#                 await client.send_message(-492221565, str(first_name) + ' сказал "' + str(one_word) + '"')
#
#             else:
#
#                 await client.send_message(-492221565, str(user) + ' сказал "'+str(one_word)+'"')
#
#             time.sleep(1)
#             if phone is None:
#                 await client.send_message(-492221565, 'Номер скрыт, слить не получится, живи пока')
#                 return
#             else:
#                 await client.send_message(-492221565, 'За тобой выехали, номер +' + str(phone) + ' уже в базах СК и МВД')
#                 return





# @app.on_message(filters.chat(eval(donor)) & filters.text)
# async def get_post(client, message):
#     username = message.chat.username
#     message_id = message.message_id
#     txt = message['text'].lower()
#     txtsplit = txt.split()
#     print(txt)
#     print(txtsplit)
#     for key in keys:
#         if key in txt:
#             counter = 0
#             for i in txtsplit:
#                 if ',' in str(i):
#                     y = str(i).find(',')
#                     print(y)
#                     try:
#                         if y == 0:
#                             message.copy(eval(own))
#                             await client.send_message(eval(own), 'По правилам грамматики, запятую нужно ставить сразу после слова, а затем пробел!')
#
#                         else:
#                             message.copy(eval(own))
#                             # await client.send_message(eval(own), message['text'])
#                             await client.send_message(eval(own), 'Запятая находится после слова ' + str(i[:-1 or None]))
#                     except Exception as e:
#                         message.copy(eval(own))
#                         await client.send_message(eval(own), 'По правилам грамматики, запятую нужно ставить сразу после слова, а затем пробел!')
#                         # return
#



